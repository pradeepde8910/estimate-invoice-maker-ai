import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.schemas.report import ReportResult

_HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TOTALS_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
_TOTALS_FONT = Font(bold=True)
_MAX_COL_WIDTH = 40


def export_excel(result: ReportResult) -> bytes:
    """
    Exports a ReportResult to Excel (.xlsx) format and returns the bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = result.report_type

    if not result.columns:
        ws.append(["No columns selected."])
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # Write header
    ws.append(result.columns)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    # Write rows
    for row_dict in result.rows:
        row_list = [row_dict.get(col, "") for col in result.columns]
        ws.append(row_list)

    # Write empty row
    ws.append([])

    # Write totals — bold and shaded so they read as a summary rather than
    # one more data row when scanning the sheet.
    totals_start_row = ws.max_row + 1
    for key, value in result.totals.items():
        totals_row = [""] * len(result.columns)
        if len(result.columns) >= 2:
            totals_row[0] = key
            totals_row[1] = value
        else:
            totals_row[0] = f"{key}: {value}"
        ws.append(totals_row)
    for row in ws.iter_rows(min_row=totals_start_row, max_row=ws.max_row):
        for cell in row:
            cell.fill = _TOTALS_FILL
            cell.font = _TOTALS_FONT

    # Approximate auto-fit: widen each column to its longest cell, capped so
    # a single long description/address doesn't blow out the whole sheet.
    for idx, col in enumerate(result.columns, start=1):
        longest = len(str(col))
        for row_dict in result.rows:
            val = row_dict.get(col, "")
            longest = max(longest, len(str(val)) if val is not None else 0)
        ws.column_dimensions[get_column_letter(idx)].width = min(longest + 2, _MAX_COL_WIDTH)

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
