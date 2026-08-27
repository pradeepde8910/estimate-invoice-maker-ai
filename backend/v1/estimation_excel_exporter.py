"""
Multi-sheet Excel export for a V1 AI estimation — the full nested pipeline
output (client info, requirements, cost breakdown, timeline, infra/license
costs, risks/assumptions) laid out as a real interactive workbook with
native (editable) Excel charts, plus a standalone timeline-only workbook.

Reads straight from `Estimation.raw_pipeline_json` (see get_document_data in
api.py) rather than the frontend's flattened JobResult shape — the raw JSON
has fields JobResult drops or gets wrong (e.g. category_breakdown has no
requirements_count; phases have an estimation_note JobResult doesn't type).
"""

from __future__ import annotations

import io
from datetime import datetime, timedelta
from typing import Any

import re

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.text import CharacterProperties

HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=12)
TOTALS_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
TOTALS_FONT = Font(bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="CBD5E1"))
MAX_COL_WIDTH = 50

# A varied, presentation-friendly palette — cycled per data point on bar
# charts (one flat color per series reads as an unfinished draft; a
# distinct color per category/phase is what an actual presentation chart
# looks like) and relied on implicitly by pie charts (Excel/openpyxl auto-
# colors pie slices per point already).
CHART_COLORS = ["4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5", "70AD47", "264478", "9E480E", "636363", "997300"]

def _table_name(label: str) -> str:
    """
    Excel table names must be unique per workbook and contain no spaces —
    slugifies the caller's label into a valid name. Deliberately stateless
    (no module-level "seen names" tracking): this module builds one
    workbook per call with a fixed, already-distinct label per sheet, and
    this function runs inside a synchronous FastAPI endpoint — which
    Starlette executes in a thread pool — so any shared mutable state here
    would be a real cross-request race, not just theoretical.
    """
    base = re.sub(r"[^A-Za-z0-9_]", "_", label).strip("_") or "Table"
    if not base[0].isalpha():
        base = f"T_{base}"
    return base


def _style_chart_title(chart, text: str, size: int = 1400) -> None:
    """
    Sets a chart's title with an explicit larger font size (openpyxl's
    plain `chart.title = "text"` works but inherits Excel's fairly small
    default chart-title size) — matters for a workbook meant to be
    presented/projected, not just opened at a desk.
    """
    chart.title = text
    try:
        run = chart.title.tx.rich.p[0].r[0]
        run.rPr = CharacterProperties(sz=size, b=True)
    except (AttributeError, IndexError):
        pass


def _safe(v, default=""):
    return v if v is not None else default


def _write_title(ws: Worksheet, text: str, row: int = 1) -> int:
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, size=14, color="1E293B")
    return row + 2


def _write_kv_block(ws: Worksheet, start_row: int, pairs: list[tuple[str, Any]]) -> int:
    """Writes a two-column label/value block (for overview-style metadata)."""
    row = start_row
    for label, value in pairs:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, color="475569")
        ws.cell(row=row, column=2, value=_safe(value))
        row += 1
    return row


def _write_table(ws: Worksheet, start_row: int, headers: list[str], rows: list[list[Any]], table_label: str | None = None) -> int:
    """
    Writes a styled header + data rows starting at start_row, and — when
    there's real data and a label is given — registers the range as a real
    Excel Table (banded rows, built-in filter/sort dropdowns on the header)
    rather than a plain styled range. That's what actually makes a workbook
    feel "interactive" to someone opening it, as opposed to a static print-
    out with colored cells. Returns the row after the last data row.
    """
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    r = start_row + 1
    if not rows:
        ws.cell(row=r, column=1, value="No data available for this estimation.")
        r += 1
    else:
        for row_values in rows:
            for c, val in enumerate(row_values, start=1):
                ws.cell(row=r, column=c, value=_safe(val))
            r += 1

    # Approximate auto-fit, capped, based on header + actual cell lengths
    for c, h in enumerate(headers, start=1):
        longest = len(str(h))
        for row_values in rows:
            if c - 1 < len(row_values) and row_values[c - 1] is not None:
                longest = max(longest, len(str(row_values[c - 1])))
        ws.column_dimensions[get_column_letter(c)].width = min(longest + 2, MAX_COL_WIDTH)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate

    if rows and table_label:
        last_col_letter = get_column_letter(len(headers))
        ref = f"{get_column_letter(1)}{start_row}:{last_col_letter}{r - 1}"
        table = Table(displayName=_table_name(table_label), ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False,
        )
        ws.add_table(table)

    return r


def _add_data_labels(chart_or_series, number_format: str = "#,##0", show_val: bool = True, show_cat: bool = False, show_pct: bool = False) -> None:
    """
    Works on either a chart (PieChart/BarChart, which alias `.dataLabels` to
    `.dLbls`) or an individual `Series` (which has ONLY `.dLbls` — no such
    alias). Always setting `.dLbls` directly is the one spelling that's
    correct for both, so callers can't hit the series case silently no-op
    (that's exactly what happened here) by using the chart-only spelling.
    """
    labels = DataLabelList()
    labels.showVal = show_val
    labels.showCatName = show_cat
    labels.showPercent = show_pct
    labels.showLegendKey = False
    labels.showSerName = False
    labels.numFmt = number_format
    chart_or_series.dLbls = labels


def _color_points(series, count: int) -> None:
    """Gives each bar its own color from the palette (cycled) instead of
    every bar in the series sharing one flat color — the single biggest
    thing that makes an openpyxl bar chart look like a real presentation
    chart rather than a spreadsheet default."""
    series.data_points = [
        DataPoint(idx=i, spPr=GraphicalProperties(solidFill=CHART_COLORS[i % len(CHART_COLORS)]))
        for i in range(count)
    ]


def _bar_chart(title: str, ws: Worksheet, cat_ref: Reference, val_ref: Reference, anchor: str, y_title: str = "", horizontal: bool = False, point_count: int = 0, number_format: str = "#,##0") -> BarChart:
    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    # The unit hint goes in the chart title itself, e.g. "Total Cost by
    # Role (INR)", rather than a separate rotated y-axis title. openpyxl/
    # Excel's auto-layout doesn't always leave enough room between a
    # rotated axis title and the tick-number column next to it — they can
    # render on top of each other — and there's no reliable way to hand-fix
    # that spacing without risking malformed chart XML. Folding the unit
    # into the title sidesteps the collision entirely rather than fighting
    # for margin, at the cost of one axis label the chart doesn't strictly
    # need (data labels on the bars already carry the real numbers).
    _style_chart_title(chart, f"{title} ({y_title})" if y_title else title)
    chart.y_axis.delete = False
    chart.y_axis.majorTickMark = "out"
    chart.y_axis.numFmt = number_format
    chart.x_axis.delete = False
    chart.style = 10
    chart.height = 10
    chart.width = 20
    chart.add_data(val_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.legend = None
    if chart.series:
        _color_points(chart.series[0], point_count or 1)
        _add_data_labels(chart.series[0], number_format=number_format)
    ws.add_chart(chart, anchor)
    return chart


def _pie_chart(title: str, ws: Worksheet, cat_ref: Reference, val_ref: Reference, anchor: str, number_format: str = "#,##0") -> PieChart:
    chart = PieChart()
    _style_chart_title(chart, title)
    chart.height = 10
    chart.width = 16
    chart.add_data(val_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    # Category name is deliberately left out of the label: the legend
    # already maps color -> category, and stacking category+value+percent
    # into one label produces long strings that overlap each other (and
    # the title) around thin slices. Value + percent is enough on the pie
    # itself; the source table nearby has the rest.
    _add_data_labels(chart, number_format=number_format, show_val=True, show_cat=False, show_pct=True)
    ws.add_chart(chart, anchor)
    return chart


# ─────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────

def _sheet_overview(wb: Workbook, data: dict, est, client) -> None:
    ws = wb.active
    ws.title = "Overview"
    analysis = data.get("analysis") or {}
    cost = data.get("cost_estimation") or {}

    row = _write_title(ws, f"Estimation Overview — {data.get('project_name') or analysis.get('project_name') or 'Untitled Project'}")

    ws.cell(row=row, column=1, value="Client").font = SUBHEADER_FONT
    row += 1
    client_pairs = []
    if client:
        client_pairs = [
            ("Company Name", client.company_name),
            ("Contact Person", client.contact_person),
            ("Email", client.email),
            ("Phone", client.phone),
            ("GSTIN", client.gstin),
            ("Billing Address", client.billing_address),
        ]
    else:
        client_info = data.get("client_info") or {}
        client_pairs = [(k.replace("_", " ").title(), v) for k, v in client_info.items()]
    row = _write_kv_block(ws, row, client_pairs) + 1

    ws.cell(row=row, column=1, value="Project").font = SUBHEADER_FONT
    row += 1
    row = _write_kv_block(ws, row, [
        ("Estimation Number", getattr(est, "estimation_number", None)),
        ("Project Type", analysis.get("project_type")),
        ("Target Audience", analysis.get("target_audience")),
        ("Status", data.get("status") or getattr(est, "status", None)),
        ("Version", data.get("version") or getattr(est, "version", None)),
        ("Generated At", data.get("generated_at")),
        ("Created At", str(getattr(est, "created_at", "")) if est else ""),
        ("Tech Stack", ", ".join(analysis.get("tech_stack_suggested") or [])),
    ]) + 1

    description = analysis.get("project_description")
    if description:
        ws.cell(row=row, column=1, value="Description").font = SUBHEADER_FONT
        row += 1
        ws.cell(row=row, column=1, value=description).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 60
        row += 2

    ws.cell(row=row, column=1, value="Financial Summary").font = SUBHEADER_FONT
    row += 1
    dev_cost = float(cost.get("total_development_cost") or 0)
    contingency = float(cost.get("contingency_amount") or 0)
    infra_6mo = float(cost.get("infrastructure_cost_monthly") or 0) * 6
    licenses_6mo = float(cost.get("third_party_licenses_monthly") or 0) * 6
    misc_cost = float(cost.get("miscellaneous_costs") or 0)
    grand_total = float(cost.get("grand_total") or (dev_cost + contingency + infra_6mo + licenses_6mo + misc_cost))

    fin_rows = [
        ("Development Cost", dev_cost),
        ("Contingency", contingency),
        ("Infrastructure (6 months)", infra_6mo),
        ("Third-Party Licenses (6 months)", licenses_6mo),
    ]
    if misc_cost:
        fin_rows.append(("Miscellaneous Costs", misc_cost))
    table_start = row
    for label, val in fin_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val).number_format = "#,##0.00"
        row += 1
    ws.cell(row=row, column=1, value="Grand Total").font = TOTALS_FONT
    ws.cell(row=row, column=2, value=grand_total).font = TOTALS_FONT
    ws.cell(row=row, column=2).number_format = "#,##0.00"
    ws.cell(row=row, column=1).fill = TOTALS_FILL
    ws.cell(row=row, column=2).fill = TOTALS_FILL
    row += 1

    ws.cell(row=row, column=1, value="Timeline (weeks)")
    ws.cell(row=row, column=2, value=float(cost.get("timeline_weeks") or 0))
    row += 2

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40

    # Grand total composition — a genuinely useful at-a-glance chart: what
    # actually drives the price (dev labor vs. contingency buffer vs.
    # recurring infra/license costs the client will keep paying).
    if any(v > 0 for _, v in fin_rows):
        cat_ref = Reference(ws, min_col=1, min_row=table_start, max_row=table_start + len(fin_rows) - 1)
        val_ref = Reference(ws, min_col=2, min_row=table_start - 1, max_row=table_start + len(fin_rows) - 1)
        _pie_chart("Grand Total Composition", ws, cat_ref, val_ref, f"D{table_start - 1}", number_format="#,##0.00")


def _sheet_timeline(wb: Workbook, data: dict, est, sheet_name: str = "Timeline") -> None:
    ws = wb.create_sheet(sheet_name)
    cost = data.get("cost_estimation") or {}
    phases = cost.get("phases") or []
    timeline_weeks = float(cost.get("timeline_weeks") or 0)

    row = _write_title(ws, "Project Timeline (Phases)")
    ws.cell(row=row, column=1, value=f"Total estimated duration: {timeline_weeks:g} weeks").font = Font(italic=True, color="64748B")
    row += 1
    ws.cell(row=row, column=1, value=(
        "Durations are relative (in weeks from project start) — the estimation pipeline doesn't fix calendar "
        "dates, so 'Start Date'/'End Date' below are computed from this workbook's generation date. Adjust "
        "the start date in cell if your actual kickoff differs; downstream cells don't recompute automatically."
    )).font = Font(italic=True, size=9, color="94A3B8")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    start_date = (getattr(est, "created_at", None) or datetime.utcnow()).date()
    headers = ["Phase", "Start Week", "End Week", "Duration (Weeks)", "Est. Start Date", "Est. End Date", "Description"]
    table_start = row
    rows = []
    elapsed = 0.0
    for phase in phases:
        duration = float(phase.get("duration_weeks") or 0)
        start_week = elapsed
        end_week = elapsed + duration
        elapsed = end_week
        rows.append([
            phase.get("name") or "Untitled Phase",
            round(start_week, 1),
            round(end_week, 1),
            round(duration, 1),
            start_date + timedelta(weeks=start_week),
            start_date + timedelta(weeks=end_week),
            phase.get("description") or phase.get("estimation_note") or "",
        ])
    end_row = _write_table(ws, table_start, headers, rows, table_label="Timeline")

    if rows:
        # Gantt-style horizontal stacked bar: an invisible "start offset"
        # series pushes each phase's visible bar to begin at its actual
        # start week, so the chart reads as a real Gantt chart rather than
        # a plain bar chart of durations.
        data_first_row = table_start + 1
        data_last_row = end_row - 1
        cat_ref = Reference(ws, min_col=1, min_row=data_first_row, max_row=data_last_row)
        start_ref = Reference(ws, min_col=2, min_row=table_start, max_row=data_last_row)
        dur_ref = Reference(ws, min_col=4, min_row=table_start, max_row=data_last_row)

        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "stacked"
        chart.overlap = 100
        _style_chart_title(chart, "Project Timeline (Gantt)")
        chart.height = max(8, 1.2 * len(rows))
        chart.width = 22
        chart.add_data(start_ref, titles_from_data=True)
        chart.add_data(dur_ref, titles_from_data=True)
        chart.set_categories(cat_ref)

        # IMPORTANT: for a BarChart, `x_axis` is always the CATEGORY axis
        # (phase names) and `y_axis` is always the VALUE axis (weeks) —
        # this does NOT swap when type="bar" (horizontal); only the bar
        # direction changes. Getting these backwards (reversing y_axis
        # instead of x_axis, as an earlier version of this did) both
        # scrambles which phase visually starts first/last AND can leave
        # the value axis' tick numbers from rendering at all.
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.y_axis.title = "Week"
        chart.y_axis.majorTickMark = "out"
        chart.y_axis.numFmt = "0"
        # Phase 1 reads top-to-bottom like a real Gantt chart instead of
        # bottom-to-top (Excel's default category order for horizontal bar
        # charts) — this reverses the CATEGORY axis, not the value axis.
        chart.x_axis.scaling.orientation = "maxMin"

        # First series (start offset) invisible; second (duration) is the
        # visible Gantt bar, one distinct color per phase with its actual
        # duration printed directly on the bar.
        chart.series[0].graphicalProperties.noFill = True
        _color_points(chart.series[1], len(rows))
        _add_data_labels(chart.series[1], number_format="0.# \"wk\"")
        chart.legend = None
        ws.add_chart(chart, f"A{end_row + 2}")

    ws.column_dimensions["G"].width = 50


def _sheet_cost_by_role(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Cost by Role")
    role_estimates = sorted(data.get("cost_estimation", {}).get("role_estimates") or [], key=lambda r: -float(r.get("total_cost") or 0))

    row = _write_title(ws, "Cost & Effort by Role")
    headers = ["Role", "Hours", "Rate / Hour", "Total Cost"]
    rows = [[r.get("role_label") or r.get("role_key"), float(r.get("hours") or 0), float(r.get("rate_per_hour") or 0), float(r.get("total_cost") or 0)] for r in role_estimates]
    end_row = _write_table(ws, row, headers, rows, table_label="CostByRole")

    if rows:
        data_first, data_last = row + 1, end_row - 1
        cat_ref = Reference(ws, min_col=1, min_row=data_first, max_row=data_last)
        cost_ref = Reference(ws, min_col=4, min_row=row, max_row=data_last)
        hours_ref = Reference(ws, min_col=2, min_row=row, max_row=data_last)
        _bar_chart("Total Cost by Role", ws, cat_ref, cost_ref, f"F{row}", y_title="INR", point_count=len(rows))
        _pie_chart("Effort Distribution (Hours) by Role", ws, cat_ref, hours_ref, f"F{row + 18}", number_format="#,##0.0")


def _sheet_cost_by_category(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Cost by Category")
    breakdown = sorted(data.get("cost_estimation", {}).get("category_breakdown") or [], key=lambda c: -float(c.get("total_cost") or 0))

    row = _write_title(ws, "Cost & Effort by Requirement Category")
    headers = ["Category", "Total Hours", "Total Cost"]
    rows = [[c.get("category"), float(c.get("total_hours") or 0), float(c.get("total_cost") or 0)] for c in breakdown]
    end_row = _write_table(ws, row, headers, rows, table_label="CostByCategory")

    if rows:
        data_first, data_last = row + 1, end_row - 1
        cat_ref = Reference(ws, min_col=1, min_row=data_first, max_row=data_last)
        cost_ref = Reference(ws, min_col=3, min_row=row, max_row=data_last)
        _pie_chart("Cost Share by Category", ws, cat_ref, cost_ref, f"E{row}")


def _sheet_requirements(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Requirements")
    analysis = data.get("analysis") or {}
    units = analysis.get("project_units") or []

    row = _write_title(ws, "Requirements (as analyzed)")
    headers = ["Unit / Phase", "Requirement", "Category", "Priority", "Complexity", "Scope Status", "Technologies", "Description"]
    rows = []
    category_counts: dict[str, int] = {}
    for unit in units:
        for req in unit.get("requirements") or []:
            category = req.get("category") or "Uncategorized"
            category_counts[category] = category_counts.get(category, 0) + 1
            rows.append([
                unit.get("label"),
                req.get("title"),
                category,
                req.get("priority"),
                req.get("complexity"),
                req.get("scope_status"),
                ", ".join(req.get("technologies") or []),
                req.get("description"),
            ])
    end_row = _write_table(ws, row, headers, rows, table_label="Requirements")

    if category_counts:
        chart_start = end_row + 2
        ws.cell(row=chart_start, column=1, value="Requirement Count by Category").font = SUBHEADER_FONT
        chart_start += 1
        sorted_counts = sorted(category_counts.items(), key=lambda kv: -kv[1])
        for cat, count in sorted_counts:
            ws.cell(row=chart_start, column=1, value=cat)
            ws.cell(row=chart_start, column=2, value=count)
            chart_start += 1
        cat_ref = Reference(ws, min_col=1, min_row=end_row + 3, max_row=chart_start - 1)
        val_ref = Reference(ws, min_col=2, min_row=end_row + 2, max_row=chart_start - 1)
        _bar_chart("Requirement Count by Category", ws, cat_ref, val_ref, f"D{end_row + 2}", point_count=len(sorted_counts), number_format="0")


def _sheet_units_and_tasks(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Units & Tasks")
    unit_estimates = data.get("cost_estimation", {}).get("unit_estimates") or []

    row = _write_title(ws, "Detailed Estimate — Units, Requirements & Tasks")
    headers = ["Unit / Phase", "Unit Hours", "Unit Cost", "Requirement", "Req. Hours", "Req. Cost", "Task", "Role", "Task Hours", "Task Cost"]
    rows = []
    unit_hours_rollup: list[tuple[str, float]] = []
    for unit in unit_estimates:
        unit_label = unit.get("label")
        unit_est = unit.get("estimate") or {}
        unit_hours_rollup.append((unit_label, float(unit_est.get("hours") or 0)))
        req_estimates = unit.get("requirement_estimates") or []
        if not req_estimates:
            rows.append([unit_label, unit_est.get("hours"), unit_est.get("cost"), "", "", "", "", "", "", ""])
            continue
        for req in req_estimates:
            tasks = req.get("implementation_tasks") or []
            if not tasks:
                rows.append([unit_label, unit_est.get("hours"), unit_est.get("cost"), req.get("title"), req.get("hours"), req.get("cost"), "", "", "", ""])
                continue
            for task in tasks:
                rows.append([
                    unit_label, unit_est.get("hours"), unit_est.get("cost"),
                    req.get("title"), req.get("hours"), req.get("cost"),
                    task.get("task"), task.get("role_label") or task.get("role_key"),
                    task.get("hours"), task.get("cost"),
                ])
    end_row = _write_table(ws, row, headers, rows, table_label="UnitsAndTasks")

    if unit_hours_rollup:
        chart_start = end_row + 2
        ws.cell(row=chart_start, column=1, value="Effort (Hours) by Unit / Phase").font = SUBHEADER_FONT
        chart_start += 1
        first_data_row = chart_start
        for label, hours in unit_hours_rollup:
            ws.cell(row=chart_start, column=1, value=label)
            ws.cell(row=chart_start, column=2, value=hours)
            chart_start += 1
        cat_ref = Reference(ws, min_col=1, min_row=first_data_row, max_row=chart_start - 1)
        val_ref = Reference(ws, min_col=2, min_row=first_data_row - 1, max_row=chart_start - 1)
        _bar_chart("Effort (Hours) by Unit / Phase", ws, cat_ref, val_ref, f"D{first_data_row - 1}", point_count=len(unit_hours_rollup), number_format="#,##0.0")


def _sheet_team(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Team Composition")
    team = data.get("cost_estimation", {}).get("team_composition") or []
    row = _write_title(ws, "Recommended Team Composition")
    headers = ["Role", "Headcount", "Billing Status", "Justification"]
    rows = [[t.get("role_key"), t.get("count"), t.get("billing_status"), t.get("justification")] for t in team]
    end_row = _write_table(ws, row, headers, rows, table_label="TeamComposition")

    if rows:
        data_first, data_last = row + 1, end_row - 1
        cat_ref = Reference(ws, min_col=1, min_row=data_first, max_row=data_last)
        count_ref = Reference(ws, min_col=2, min_row=row, max_row=data_last)
        _bar_chart("Headcount by Role", ws, cat_ref, count_ref, f"F{row}", point_count=len(rows), number_format="0")


def _sheet_infrastructure(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Infrastructure & Licenses")
    items = (data.get("web_search_data") or {}).get("items") or []

    row = _write_title(ws, "Infrastructure & Third-Party Cost Estimates")
    headers = ["Service", "Category", "Cost Type", "Tier", "Quantity", "Unit", "Billing Model", "Monthly Cost (INR)", "Cost Basis"]
    rows = []
    for item in items:
        cfg = item.get("configuration") or {}
        rows.append([
            item.get("service_name"), item.get("service_category"), item.get("cost_type"),
            cfg.get("tier"), cfg.get("quantity"), cfg.get("unit"),
            item.get("billing_model"), item.get("monthly_cost_inr"), item.get("cost_basis"),
        ])
    end_row = _write_table(ws, row, headers, rows, table_label="Infrastructure")

    costed = [(it.get("service_name") or "Unnamed", float(it.get("monthly_cost_inr") or 0)) for it in items if it.get("monthly_cost_inr")]
    if costed:
        chart_start = end_row + 2
        ws.cell(row=chart_start, column=1, value="Monthly Cost by Service").font = SUBHEADER_FONT
        chart_start += 1
        first = chart_start
        sorted_costed = sorted(costed, key=lambda x: -x[1])
        for name, cost in sorted_costed:
            ws.cell(row=chart_start, column=1, value=name)
            ws.cell(row=chart_start, column=2, value=cost)
            chart_start += 1
        cat_ref = Reference(ws, min_col=1, min_row=first, max_row=chart_start - 1)
        val_ref = Reference(ws, min_col=2, min_row=first - 1, max_row=chart_start - 1)
        _bar_chart("Monthly Cost by Service", ws, cat_ref, val_ref, f"D{first - 1}", y_title="INR / month", point_count=len(sorted_costed))


def _sheet_risks(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Risks & Assumptions")
    analysis = data.get("analysis") or {}
    cost = data.get("cost_estimation") or {}

    row = _write_title(ws, "Risks, Assumptions & Scope Notes")

    def _list_block(title: str, items: list[str], r: int) -> int:
        ws.cell(row=r, column=1, value=title).font = SUBHEADER_FONT
        r += 1
        if not items:
            ws.cell(row=r, column=1, value="None recorded.")
            r += 1
        else:
            for item in items:
                ws.cell(row=r, column=1, value=f"• {item}")
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
                r += 1
        return r + 1

    row = _list_block("Assumptions", analysis.get("assumptions") or [], row)
    row = _list_block("Estimation Assumptions", cost.get("estimation_assumptions") or [], row)
    row = _list_block("Risks", analysis.get("risks") or [], row)
    row = _list_block("Out of Scope", analysis.get("out_of_scope") or [], row)
    if cost.get("contingency_rationale"):
        row = _list_block("Contingency Rationale", [cost["contingency_rationale"]], row)

    ws.column_dimensions["A"].width = 100


# ─────────────────────────────────────────────────────────────────────────
# Public entry points
# ─────────────────────────────────────────────────────────────────────────

def build_full_workbook(data: dict, est=None, client=None) -> bytes:
    """
    Builds the complete multi-sheet estimation workbook: overview, timeline
    (with Gantt chart), cost breakdowns by role/category (with charts),
    requirements, unit/task-level detail, team composition, infra/license
    costs, and risks/assumptions.
    """
    wb = Workbook()
    _sheet_overview(wb, data, est, client)
    _sheet_timeline(wb, data, est)
    _sheet_cost_by_role(wb, data)
    _sheet_cost_by_category(wb, data)
    _sheet_requirements(wb, data)
    _sheet_units_and_tasks(wb, data)
    _sheet_team(wb, data)
    _sheet_infrastructure(wb, data)
    _sheet_risks(wb, data)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_timeline_workbook(data: dict, est=None) -> bytes:
    """Standalone single-focus workbook: just the phase timeline + Gantt chart."""
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_timeline(wb, data, est, sheet_name="Timeline")
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
